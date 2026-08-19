"""Excel menu import: template, parse, and in-memory job store.

Jobs live in process memory (single API worker). A new DB session is opened
inside the background task so the request session is never reused.
"""

from __future__ import annotations

import asyncio
import csv
import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Literal
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.core.database import SessionFactory
from app.core.errors import NotFoundError, ValidationError
from app.core.logging import get_logger
from app.schemas.menu import ImportJobResponse
from app.services.menu import MenuService

logger = get_logger("aahaar.menu_import")

MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 500
JOB_TTL_SECONDS = 2 * 60 * 60
DEFAULT_CATEGORY = "Main Course"
TEMPLATE_FILENAME = "aahaar-menu-example.xlsx"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ImportStatus = Literal["pending", "running", "done", "failed"]

_HEADER_NAME = frozenset({"dish", "dish name", "name", "item", "item name"})
_HEADER_CATEGORY = frozenset({"category"})
_HEADER_PRICE = frozenset(
    {
        "price",
        "price (inr)",
        "price (rupees)",
        "price inr",
        "inr",
        "rupees",
        "amount",
    }
)
_CATEGORY_ALIASES = {
    "starter": "Starters",
    "starters": "Starters",
    "appetizer": "Starters",
    "appetisers": "Starters",
    "appetizers": "Starters",
    "main": "Main Course",
    "main course": "Main Course",
    "mains": "Main Course",
    "entree": "Main Course",
    "entrée": "Main Course",
    "drink": "Drinks",
    "drinks": "Drinks",
    "beverage": "Drinks",
    "beverages": "Drinks",
    "sweet": "Sweets",
    "sweets": "Sweets",
    "dessert": "Sweets",
    "desserts": "Sweets",
}


@dataclass(frozen=True)
class ImportRow:
    name: str
    category: str
    price: Decimal
    description: str | None = None
    is_vegetarian: bool = True


@dataclass
class ImportJob:
    id: uuid.UUID
    restaurant_id: uuid.UUID
    status: ImportStatus = "pending"
    created: int = 0
    skipped: int = 0
    error: str | None = None
    started_at: datetime = field(default_factory=lambda: datetime.now(UTC))

    def to_response(self) -> ImportJobResponse:
        return ImportJobResponse(
            job_id=self.id,
            status=self.status,
            created=self.created,
            skipped=self.skipped,
            error=self.error,
        )


_jobs: dict[uuid.UUID, ImportJob] = {}
_jobs_lock = asyncio.Lock()
_tasks: set[asyncio.Task[None]] = set()


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Menu"
    sheet.append(["Dish name", "Category", "Price (INR)"])
    bold = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold
    sheet.append(["Butter chicken", "Main Course", 280])
    sheet.append(["Paneer tikka", "Starters", 220])
    sheet.append(["Mango lassi", "Drinks", 90])
    sheet.append(["Gulab jamun", "Sweets", 80])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 14
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_xlsx(payload: bytes) -> list[ImportRow]:
    try:
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    except (BadZipFile, OSError, KeyError, ValueError) as exc:
        raise ValidationError(
            "Could not read that Excel file. Download the example and try again."
        ) from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise ValidationError("The Excel file has no sheet.")

        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            raise ValidationError("The Excel file is empty.")

        name_idx, category_idx, price_idx = _header_indexes(header_row)
        parsed: list[ImportRow] = []
        for raw in rows:
            if len(parsed) >= MAX_IMPORT_ROWS:
                break
            if raw is None or _row_empty(raw):
                continue
            name = _cell_text(_at(raw, name_idx))
            if not name:
                continue
            category = resolve_category_name(_cell_text(_at(raw, category_idx)))
            price = parse_price(_at(raw, price_idx))
            if price is None:
                continue
            parsed.append(ImportRow(name=name[:160], category=category, price=price))
        return parsed
    finally:
        workbook.close()


def parse_csv(payload: bytes) -> list[ImportRow]:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValidationError("Could not read that CSV. Save it as UTF-8 and try again.") from exc

    reader = csv.reader(io.StringIO(text))
    header_row = next(reader, None)
    if not header_row:
        raise ValidationError("The CSV file is empty.")

    name_idx, category_idx, price_idx = _header_indexes(tuple(header_row))
    parsed: list[ImportRow] = []
    for raw in reader:
        if len(parsed) >= MAX_IMPORT_ROWS:
            break
        row = tuple(raw)
        if _row_empty(row):
            continue
        name = _cell_text(_at(row, name_idx))
        if not name:
            continue
        category = resolve_category_name(_cell_text(_at(row, category_idx)))
        price = parse_price(_at(row, price_idx))
        if price is None:
            continue
        parsed.append(ImportRow(name=name[:160], category=category, price=price))
    return parsed


def resolve_category_name(raw: str) -> str:
    key = " ".join(raw.split()).casefold()
    if not key:
        return DEFAULT_CATEGORY
    return _CATEGORY_ALIASES.get(key, raw.strip())[:120]


def parse_price(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        amount = value
    elif isinstance(value, int):
        amount = Decimal(value)
    elif isinstance(value, float):
        amount = Decimal(str(value))
    else:
        text = re.sub(r"[₹,\s]", "", str(value))
        text = re.sub(r"(?i)^rs\.?", "", text)
        try:
            amount = Decimal(text)
        except InvalidOperation:
            return None
    if amount < 0:
        return None
    return amount.quantize(Decimal("0.01"))


def validate_upload(filename: str | None, size: int | None) -> None:
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        raise ValidationError("Save the file as .xlsx and try again.")
    if name and not name.endswith(".xlsx"):
        raise ValidationError("Upload an Excel file (.xlsx).")
    if size is not None and size > MAX_IMPORT_BYTES:
        raise ValidationError("That file is too large. Keep it under 5 MB.")


async def start_import(restaurant_id: uuid.UUID, payload: bytes) -> ImportJob:
    if len(payload) > MAX_IMPORT_BYTES:
        raise ValidationError("That file is too large. Keep it under 5 MB.")
    if not payload:
        raise ValidationError("The Excel file is empty.")

    job = ImportJob(id=uuid.uuid4(), restaurant_id=restaurant_id)
    async with _jobs_lock:
        _prune_jobs_locked()
        _jobs[job.id] = job

    task = asyncio.create_task(_run_import(job.id, restaurant_id, payload))
    _tasks.add(task)
    task.add_done_callback(_tasks.discard)
    return job


async def get_job(job_id: uuid.UUID, restaurant_id: uuid.UUID) -> ImportJob:
    async with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None or job.restaurant_id != restaurant_id:
        raise NotFoundError("Import job not found.")
    return job


def _header_indexes(header_row: tuple[object, ...]) -> tuple[int, int, int]:
    name_idx = category_idx = price_idx = None
    for index, cell in enumerate(header_row):
        key = _cell_text(cell).lower()
        if key in _HEADER_NAME and name_idx is None:
            name_idx = index
        elif key in _HEADER_CATEGORY and category_idx is None:
            category_idx = index
        elif key in _HEADER_PRICE and price_idx is None:
            price_idx = index
    if name_idx is None or price_idx is None:
        raise ValidationError(
            "Use columns Dish name, Category, and Price (INR). Download the example if you are unsure."
        )
    if category_idx is None:
        category_idx = -1
    return name_idx, category_idx, price_idx


def _at(row: tuple[object, ...], index: int) -> object:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_empty(row: tuple[object, ...]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _prune_jobs_locked() -> None:
    now = datetime.now(UTC)
    stale = [
        job_id
        for job_id, job in _jobs.items()
        if (now - job.started_at).total_seconds() > JOB_TTL_SECONDS
    ]
    for job_id in stale:
        _jobs.pop(job_id, None)


async def _run_import(job_id: uuid.UUID, restaurant_id: uuid.UUID, payload: bytes) -> None:
    async with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None:
            return
        job.status = "running"

    try:
        rows = parse_xlsx(payload)
        if not rows:
            raise ValidationError(
                "No dishes found in that file. Use Dish name, Category, and Price (INR)."
            )
        async with SessionFactory() as session:
            created, skipped = await MenuService(session).import_dishes(restaurant_id, rows)
        async with _jobs_lock:
            job = _jobs.get(job_id)
            if job is None:
                return
            job.created = created
            job.skipped = skipped
            job.status = "done"
    except ValidationError as exc:
        async with _jobs_lock:
            job = _jobs.get(job_id)
            if job is None:
                return
            job.status = "failed"
            job.error = exc.message
    except Exception:
        logger.exception("Menu import failed for restaurant %s job %s", restaurant_id, job_id)
        async with _jobs_lock:
            job = _jobs.get(job_id)
            if job is None:
                return
            job.status = "failed"
            job.error = "Could not import that file. Check the example and try again."
